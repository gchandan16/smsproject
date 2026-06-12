# backend/routers/finance_reports.py
# ─────────────────────────────────────────────────────────────
# Finance Reports — collection, outstanding, transport fees,
# daily cash book, ledger, all exportable to Excel + PDF
# Registered at /api/finance-reports
# ─────────────────────────────────────────────────────────────
from uuid import UUID
from datetime import date
from typing import Optional
from io import BytesIO

from fastapi import APIRouter, Depends, Query, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import text

from database import get_db
from routers.auth import get_current_user
from models.user import User

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import mm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

router = APIRouter()


def tid(cu): return str(cu.tenant_id)


# ─────────────────────────────────────────────────────────────
#  SHARED HELPERS — Excel / PDF builders
# ─────────────────────────────────────────────────────────────
HEADER_FILL  = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
HEADER_FONT  = Font(color="FFFFFF", bold=True, size=11)
TOTAL_FILL   = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")
TOTAL_FONT   = Font(bold=True)
THIN_BORDER  = Border(*[Side(style="thin", color="D1D5DB")] * 4)


def build_excel(title: str, columns: list, rows: list, totals: dict = None, subtitle: str = "") -> BytesIO:
    """
    columns: [{ "key": "field", "label": "Display Name", "width": 15, "fmt": "currency"|"text"|"number" }]
    rows:    list of dicts
    totals:  optional dict { "label": "Total", values: {key: total} }
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title[:31]

    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
    cell = ws.cell(row=1, column=1, value=title)
    cell.font = Font(bold=True, size=14, color="1E3A5F")
    cell.alignment = Alignment(horizontal="center")

    row_offset = 2
    if subtitle:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(columns))
        c2 = ws.cell(row=2, column=1, value=subtitle)
        c2.font = Font(italic=True, size=10, color="64748B")
        c2.alignment = Alignment(horizontal="center")
        row_offset = 3

    header_row = row_offset + 1

    # Header
    for ci, col in enumerate(columns, start=1):
        c = ws.cell(row=header_row, column=ci, value=col["label"])
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = THIN_BORDER
        ws.column_dimensions[c.column_letter].width = col.get("width", 15)

    # Data rows
    for ri, row in enumerate(rows, start=header_row + 1):
        for ci, col in enumerate(columns, start=1):
            val = row.get(col["key"], "")
            c = ws.cell(row=ri, column=ci, value=val)
            c.border = THIN_BORDER
            if col.get("fmt") == "currency":
                c.number_format = '"₹"#,##0.00'
            elif col.get("fmt") == "number":
                c.number_format = '#,##0'
            c.alignment = Alignment(horizontal="right" if col.get("fmt") in ("currency","number") else "left")

    # Totals row
    if totals:
        tr = header_row + len(rows) + 1
        for ci, col in enumerate(columns, start=1):
            key = col["key"]
            if ci == 1:
                c = ws.cell(row=tr, column=1, value=totals.get("label", "Total"))
            elif key in totals.get("values", {}):
                c = ws.cell(row=tr, column=ci, value=totals["values"][key])
                if col.get("fmt") == "currency":
                    c.number_format = '"₹"#,##0.00'
            else:
                c = ws.cell(row=tr, column=ci, value="")
            c.font = TOTAL_FONT
            c.fill = TOTAL_FILL
            c.border = THIN_BORDER
            c.alignment = Alignment(horizontal="right" if ci > 1 else "left")

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def build_pdf(title: str, columns: list, rows: list, totals: dict = None,
               subtitle: str = "", landscape_mode: bool = True) -> BytesIO:
    buf = BytesIO()
    pagesize = landscape(A4) if landscape_mode else A4
    doc = SimpleDocTemplate(
        buf, pagesize=pagesize,
        leftMargin=15*mm, rightMargin=15*mm, topMargin=15*mm, bottomMargin=15*mm
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("title", parent=styles["Heading1"],
        fontSize=16, textColor=colors.HexColor("#1E3A5F"), alignment=1, spaceAfter=4)
    sub_style = ParagraphStyle("sub", parent=styles["Normal"],
        fontSize=9, textColor=colors.HexColor("#64748B"), alignment=1, spaceAfter=10)

    elements = [Paragraph(title, title_style)]
    if subtitle:
        elements.append(Paragraph(subtitle, sub_style))
    elements.append(Spacer(1, 6))

    # Build table data
    header = [c["label"] for c in columns]
    data = [header]
    for row in rows:
        data_row = []
        for c in columns:
            val = row.get(c["key"], "")
            if c.get("fmt") == "currency" and val != "":
                val = f"Rs. {float(val):,.2f}"
            data_row.append(str(val))
        data.append(data_row)

    if totals:
        total_row = []
        for i, c in enumerate(columns):
            if i == 0:
                total_row.append(totals.get("label", "Total"))
            elif c["key"] in totals.get("values", {}):
                v = totals["values"][c["key"]]
                if c.get("fmt") == "currency":
                    total_row.append(f"Rs. {float(v):,.2f}")
                else:
                    total_row.append(str(v))
            else:
                total_row.append("")
        data.append(total_row)

    col_widths = [c.get("pdf_width", None) for c in columns]
    table = Table(data, colWidths=col_widths, repeatRows=1)

    style_cmds = [
        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#1E3A5F")),
        ("TEXTCOLOR",  (0,0), (-1,0), colors.white),
        ("FONTNAME",   (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE",   (0,0), (-1,-1), 8),
        ("GRID",       (0,0), (-1,-1), 0.5, colors.HexColor("#D1D5DB")),
        ("ALIGN",      (0,0), (-1,0), "CENTER"),
        ("VALIGN",     (0,0), (-1,-1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0,1), (-1,-2 if totals else -1), [colors.white, colors.HexColor("#F8FAFC")]),
    ]
    # Right-align numeric columns
    for i, c in enumerate(columns):
        if c.get("fmt") in ("currency", "number"):
            style_cmds.append(("ALIGN", (i,1), (i,-1), "RIGHT"))

    if totals:
        style_cmds += [
            ("BACKGROUND", (0,-1), (-1,-1), colors.HexColor("#E8F0FE")),
            ("FONTNAME",   (0,-1), (-1,-1), "Helvetica-Bold"),
        ]

    table.setStyle(TableStyle(style_cmds))
    elements.append(table)

    # Footer
    elements.append(Spacer(1, 10))
    footer = Paragraph(
        f"<i>Generated on {date.today().strftime('%d %b %Y')} — School Management System</i>",
        ParagraphStyle("footer", parent=styles["Normal"], fontSize=8,
                       textColor=colors.HexColor("#94A3B8"), alignment=1)
    )
    elements.append(footer)

    doc.build(elements)
    buf.seek(0)
    return buf


def excel_response(buf: BytesIO, filename: str):
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}.xlsx"'}
    )


def pdf_response(buf: BytesIO, filename: str):
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}.pdf"'}
    )


# ═══════════════════════════════════════════════════════════════
#  REPORT 1 — FEE COLLECTION REPORT (school fees, all categories)
# ═══════════════════════════════════════════════════════════════
def _fee_collection_data(db, cu, from_date, to_date, grade_id=None, category=None):
    params = {"tid": tid(cu), "from": from_date, "to": to_date}
    extra = ""
    if grade_id:
        extra += " AND sec.grade_id = :grade_id"
        params["grade_id"] = str(grade_id)
    if category:
        extra += " AND fc.name = :category"
        params["category"] = category

    rows = db.execute(text(f"""
        SELECT
            fp.payment_date,
            fp.receipt_no,
            fp.method,
            fp.amount,
            s.first_name, s.last_name, s.admission_no,
            g.name AS grade_name, sec.name AS section_name,
            fc.name AS fee_category,
            fi.invoice_no
        FROM fee_payments fp
        JOIN fee_invoices fi      ON fi.id = fp.invoice_id
        JOIN students s           ON s.id  = fi.student_id
        LEFT JOIN student_enrollments se ON se.student_id=s.id AND se.academic_year_id=fi.academic_year_id
        LEFT JOIN sections sec    ON sec.id = se.section_id
        LEFT JOIN grades g        ON g.id   = sec.grade_id
        LEFT JOIN fee_invoice_items fii ON fii.invoice_id = fi.id
        LEFT JOIN fee_categories fc  ON fc.id = fii.fee_category_id
        WHERE fp.tenant_id = :tid
          AND fp.payment_date BETWEEN :from AND :to {extra}
        ORDER BY fp.payment_date, s.first_name
    """), params).fetchall()

    # Deduplicate: a payment can join to multiple invoice_items (one per category)
    seen = set()
    data = []
    for r in rows:
        key = (str(r.payment_date), r.receipt_no, r.first_name, float(r.amount))
        if key in seen:
            continue
        seen.add(key)
        data.append({
            "date":         str(r.payment_date),
            "receipt_no":   r.receipt_no or "—",
            "student_name": f"{r.first_name} {r.last_name or ''}".strip(),
            "admission_no": r.admission_no,
            "class":        f"{r.grade_name or ''} {r.section_name or ''}".strip() or "—",
            "category":     r.fee_category or "General",
            "method":       (r.method or "").title(),
            "amount":       float(r.amount),
        })
    total = sum(d["amount"] for d in data)
    return data, total


@router.get("/fees/collection")
def fee_collection_report(
    from_date: date = Query(...),
    to_date:   date = Query(...),
    grade_id:  Optional[UUID] = Query(None),
    category:  Optional[str]  = Query(None),
    db: Session = Depends(get_db),
    cu: User    = Depends(get_current_user),
):
    data, total = _fee_collection_data(db, cu, from_date, to_date, grade_id, category)
    by_method = {}
    by_category = {}
    for d in data:
        by_method[d["method"]] = by_method.get(d["method"], 0) + d["amount"]
        by_category[d["category"]] = by_category.get(d["category"], 0) + d["amount"]

    return {
        "period": {"from": str(from_date), "to": str(to_date)},
        "total_collected": total,
        "transaction_count": len(data),
        "by_method": by_method,
        "by_category": by_category,
        "transactions": data,
    }


@router.get("/fees/collection/export")
def export_fee_collection(
    fmt: str = Query("excel", regex="^(excel|pdf)$"),
    from_date: date = Query(...),
    to_date:   date = Query(...),
    grade_id:  Optional[UUID] = Query(None),
    category:  Optional[str]  = Query(None),
    db: Session = Depends(get_db),
    cu: User    = Depends(get_current_user),
):
    data, total = _fee_collection_data(db, cu, from_date, to_date, grade_id, category)

    columns = [
        {"key":"date",         "label":"Date",        "width":12, "pdf_width":55},
        {"key":"receipt_no",   "label":"Receipt No",  "width":14, "pdf_width":65},
        {"key":"student_name", "label":"Student",     "width":22, "pdf_width":110},
        {"key":"admission_no", "label":"Admission No","width":14, "pdf_width":65},
        {"key":"class",        "label":"Class",       "width":12, "pdf_width":55},
        {"key":"category",     "label":"Fee Type",    "width":15, "pdf_width":65},
        {"key":"method",       "label":"Method",      "width":12, "pdf_width":55},
        {"key":"amount",       "label":"Amount",      "width":14, "fmt":"currency", "pdf_width":65},
    ]
    totals = {"label": "Total", "values": {"amount": total}}
    subtitle = f"Period: {from_date} to {to_date}  |  Transactions: {len(data)}"

    if fmt == "excel":
        buf = build_excel("Fee Collection Report", columns, data, totals, subtitle)
        return excel_response(buf, f"fee_collection_{from_date}_to_{to_date}")
    else:
        buf = build_pdf("Fee Collection Report", columns, data, totals, subtitle)
        return pdf_response(buf, f"fee_collection_{from_date}_to_{to_date}")


# ═══════════════════════════════════════════════════════════════
#  REPORT 2 — TRANSPORT (BUS) FEE COLLECTION
# ═══════════════════════════════════════════════════════════════
def _transport_fee_data(db, cu, from_date, to_date, route_id=None):
    params = {"tid": tid(cu), "from": from_date, "to": to_date}
    extra = ""
    if route_id:
        extra += " AND tr.id = :route_id"
        params["route_id"] = str(route_id)

    # Transport fees come through fee_payments where fee_category = 'Transport'
    rows = db.execute(text(f"""
        SELECT
            fp.payment_date,
            fp.receipt_no,
            fp.method,
            fp.amount,
            s.first_name, s.last_name, s.admission_no,
            g.name AS grade_name, sec.name AS section_name,
            tr.name AS route_name, tr.route_no,
            ts.name AS stop_name
        FROM fee_payments fp
        JOIN fee_invoices fi ON fi.id = fp.invoice_id
        JOIN students s      ON s.id  = fi.student_id
        JOIN fee_invoice_items fii ON fii.invoice_id = fi.id
        JOIN fee_categories fc     ON fc.id  = fii.fee_category_id
        LEFT JOIN student_enrollments se ON se.student_id=s.id AND se.academic_year_id=fi.academic_year_id
        LEFT JOIN sections sec ON sec.id = se.section_id
        LEFT JOIN grades g     ON g.id   = sec.grade_id
        LEFT JOIN student_transport st ON st.student_id = s.id AND st.is_active=true
        LEFT JOIN transport_routes tr  ON tr.id = st.route_id
        LEFT JOIN transport_stops ts   ON ts.id = st.stop_id
        WHERE fp.tenant_id = :tid
          AND fc.name ILIKE '%transport%'
          AND fp.payment_date BETWEEN :from AND :to {extra}
        ORDER BY fp.payment_date, s.first_name
    """), params).fetchall()

    seen = set()
    data = []
    for r in rows:
        key = (str(r.payment_date), r.receipt_no, r.first_name, float(r.amount))
        if key in seen:
            continue
        seen.add(key)
        data.append({
            "date":         str(r.payment_date),
            "receipt_no":   r.receipt_no or "—",
            "student_name": f"{r.first_name} {r.last_name or ''}".strip(),
            "admission_no": r.admission_no,
            "class":        f"{r.grade_name or ''} {r.section_name or ''}".strip() or "—",
            "route":        f"{r.route_no or ''} {r.route_name or ''}".strip() or "—",
            "stop":         r.stop_name or "—",
            "method":       (r.method or "").title(),
            "amount":       float(r.amount),
        })
    total = sum(d["amount"] for d in data)
    return data, total


@router.get("/fees/transport")
def transport_fee_report(
    from_date: date = Query(...),
    to_date:   date = Query(...),
    route_id:  Optional[UUID] = Query(None),
    db: Session = Depends(get_db),
    cu: User    = Depends(get_current_user),
):
    data, total = _transport_fee_data(db, cu, from_date, to_date, route_id)
    by_route = {}
    for d in data:
        by_route[d["route"]] = by_route.get(d["route"], 0) + d["amount"]

    return {
        "period": {"from": str(from_date), "to": str(to_date)},
        "total_collected": total,
        "transaction_count": len(data),
        "by_route": by_route,
        "transactions": data,
    }


@router.get("/fees/transport/export")
def export_transport_fee(
    fmt: str = Query("excel", regex="^(excel|pdf)$"),
    from_date: date = Query(...),
    to_date:   date = Query(...),
    route_id:  Optional[UUID] = Query(None),
    db: Session = Depends(get_db),
    cu: User    = Depends(get_current_user),
):
    data, total = _transport_fee_data(db, cu, from_date, to_date, route_id)

    columns = [
        {"key":"date",         "label":"Date",        "width":12, "pdf_width":55},
        {"key":"receipt_no",   "label":"Receipt No",  "width":14, "pdf_width":65},
        {"key":"student_name", "label":"Student",     "width":22, "pdf_width":120},
        {"key":"admission_no", "label":"Admission No","width":14, "pdf_width":65},
        {"key":"class",        "label":"Class",       "width":12, "pdf_width":55},
        {"key":"route",        "label":"Route",       "width":18, "pdf_width":85},
        {"key":"stop",         "label":"Stop",        "width":15, "pdf_width":70},
        {"key":"method",       "label":"Method",      "width":12, "pdf_width":55},
        {"key":"amount",       "label":"Amount",      "width":14, "fmt":"currency", "pdf_width":65},
    ]
    totals = {"label": "Total", "values": {"amount": total}}
    subtitle = f"Period: {from_date} to {to_date}  |  Transactions: {len(data)}"

    if fmt == "excel":
        buf = build_excel("Transport Fee Collection", columns, data, totals, subtitle)
        return excel_response(buf, f"transport_fees_{from_date}_to_{to_date}")
    else:
        buf = build_pdf("Transport Fee Collection", columns, data, totals, subtitle)
        return pdf_response(buf, f"transport_fees_{from_date}_to_{to_date}")


# ═══════════════════════════════════════════════════════════════
#  REPORT 3 — OUTSTANDING FEES (all categories, defaulters list)
# ═══════════════════════════════════════════════════════════════
def _outstanding_data(db, cu, academic_year_id, grade_id=None, min_amount=0):
    params = {"tid": tid(cu), "yr": str(academic_year_id), "min_amt": min_amount}
    extra = ""
    if grade_id:
        extra += " AND sec.grade_id = :grade_id"
        params["grade_id"] = str(grade_id)

    rows = db.execute(text(f"""
        SELECT
            s.first_name, s.last_name, s.admission_no,
            g.name AS grade_name, sec.name AS section_name,
            COUNT(fi.id)::int AS invoice_count,
            COALESCE(SUM(fi.total_amount),0)::float AS total_billed,
            COALESCE(SUM(fi.paid_amount),0)::float  AS total_paid,
            COALESCE(SUM(fi.balance),0)::float      AS outstanding,
            MAX(fi.due_date) AS latest_due,
            BOOL_OR(fi.status='overdue') AS has_overdue,
            STRING_AGG(DISTINCT fi.invoice_no, ', ') AS invoice_nos
        FROM students s
        JOIN student_enrollments se ON se.student_id=s.id AND se.academic_year_id=:yr AND se.status='active'
        JOIN sections sec ON sec.id = se.section_id
        JOIN grades g     ON g.id   = sec.grade_id
        JOIN fee_invoices fi ON fi.student_id=s.id AND fi.academic_year_id=:yr AND fi.status NOT IN ('paid','cancelled')
        WHERE s.tenant_id=:tid {extra}
        GROUP BY s.id, s.first_name, s.last_name, s.admission_no, g.name, sec.name, g.order_no
        HAVING COALESCE(SUM(fi.balance),0) > :min_amt
        ORDER BY outstanding DESC, g.order_no, s.first_name
    """), params).fetchall()

    data = [
        {
            "student_name": f"{r.first_name} {r.last_name or ''}".strip(),
            "admission_no": r.admission_no,
            "class":        f"{r.grade_name} {r.section_name}",
            "invoices":     r.invoice_nos or "—",
            "total_billed": r.total_billed,
            "total_paid":   r.total_paid,
            "outstanding":  r.outstanding,
            "due_date":     str(r.latest_due) if r.latest_due else "—",
            "status":       "OVERDUE" if r.has_overdue else "Pending",
        }
        for r in rows
    ]
    total = sum(d["outstanding"] for d in data)
    return data, total


@router.get("/fees/outstanding")
def outstanding_report(
    academic_year_id: UUID = Query(...),
    grade_id: Optional[UUID] = Query(None),
    min_amount: float = Query(0),
    db: Session = Depends(get_db),
    cu: User    = Depends(get_current_user),
):
    data, total = _outstanding_data(db, cu, academic_year_id, grade_id, min_amount)
    return {
        "student_count": len(data),
        "total_outstanding": total,
        "students": data,
    }


@router.get("/fees/outstanding/export")
def export_outstanding(
    fmt: str = Query("excel", regex="^(excel|pdf)$"),
    academic_year_id: UUID = Query(...),
    grade_id: Optional[UUID] = Query(None),
    min_amount: float = Query(0),
    db: Session = Depends(get_db),
    cu: User    = Depends(get_current_user),
):
    data, total = _outstanding_data(db, cu, academic_year_id, grade_id, min_amount)

    columns = [
        {"key":"student_name", "label":"Student",      "width":22, "pdf_width":110},
        {"key":"admission_no", "label":"Admission No", "width":14, "pdf_width":65},
        {"key":"class",        "label":"Class",        "width":12, "pdf_width":55},
        {"key":"invoices",     "label":"Invoice No(s)","width":20, "pdf_width":90},
        {"key":"total_billed", "label":"Billed",       "width":12, "fmt":"currency", "pdf_width":60},
        {"key":"total_paid",   "label":"Paid",         "width":12, "fmt":"currency", "pdf_width":60},
        {"key":"outstanding",  "label":"Outstanding",  "width":14, "fmt":"currency", "pdf_width":65},
        {"key":"due_date",     "label":"Due Date",     "width":12, "pdf_width":55},
        {"key":"status",       "label":"Status",       "width":10, "pdf_width":50},
    ]
    totals = {"label": "Total", "values": {"outstanding": total}}
    subtitle = f"Students with dues: {len(data)}  |  Total Outstanding: Rs. {total:,.2f}"

    if fmt == "excel":
        buf = build_excel("Outstanding Fees Report", columns, data, totals, subtitle)
        return excel_response(buf, "outstanding_fees")
    else:
        buf = build_pdf("Outstanding Fees Report", columns, data, totals, subtitle)
        return pdf_response(buf, "outstanding_fees")


# ═══════════════════════════════════════════════════════════════
#  REPORT 4 — DAILY CASH BOOK (day-wise, method-wise summary)
# ═══════════════════════════════════════════════════════════════
def _daily_cashbook_data(db, cu, from_date, to_date):
    rows = db.execute(text("""
        SELECT
            fp.payment_date,
            fp.method,
            COUNT(*)::int AS txn_count,
            COALESCE(SUM(fp.amount),0)::float AS total
        FROM fee_payments fp
        WHERE fp.tenant_id=:tid
          AND fp.payment_date BETWEEN :from AND :to
        GROUP BY fp.payment_date, fp.method
        ORDER BY fp.payment_date
    """), {"tid": tid(cu), "from": from_date, "to": to_date}).fetchall()

    by_date = {}
    for r in rows:
        d = str(r.payment_date)
        if d not in by_date:
            by_date[d] = {"date": d, "cash": 0, "card": 0, "online": 0,
                          "cheque": 0, "upi": 0, "total": 0, "txn_count": 0}
        method = (r.method or "cash").lower()
        if method not in by_date[d]:
            by_date[d][method] = 0
        by_date[d][method] += r.total
        by_date[d]["total"] += r.total
        by_date[d]["txn_count"] += r.txn_count

    data = list(by_date.values())
    total = sum(d["total"] for d in data)
    return data, total


@router.get("/fees/daily-cashbook")
def daily_cashbook_report(
    from_date: date = Query(...),
    to_date:   date = Query(...),
    db: Session = Depends(get_db),
    cu: User    = Depends(get_current_user),
):
    data, total = _daily_cashbook_data(db, cu, from_date, to_date)
    return {
        "period": {"from": str(from_date), "to": str(to_date)},
        "grand_total": total,
        "days": data,
    }


@router.get("/fees/daily-cashbook/export")
def export_daily_cashbook(
    fmt: str = Query("excel", regex="^(excel|pdf)$"),
    from_date: date = Query(...),
    to_date:   date = Query(...),
    db: Session = Depends(get_db),
    cu: User    = Depends(get_current_user),
):
    data, total = _daily_cashbook_data(db, cu, from_date, to_date)

    columns = [
        {"key":"date",      "label":"Date",       "width":12, "pdf_width":60},
        {"key":"cash",      "label":"Cash",       "width":12, "fmt":"currency", "pdf_width":65},
        {"key":"card",      "label":"Card",       "width":12, "fmt":"currency", "pdf_width":65},
        {"key":"online",    "label":"Online",     "width":12, "fmt":"currency", "pdf_width":65},
        {"key":"upi",       "label":"UPI",        "width":12, "fmt":"currency", "pdf_width":65},
        {"key":"cheque",    "label":"Cheque",     "width":12, "fmt":"currency", "pdf_width":65},
        {"key":"txn_count", "label":"Txns",       "width":8,  "fmt":"number",   "pdf_width":45},
        {"key":"total",     "label":"Total",      "width":14, "fmt":"currency", "pdf_width":70},
    ]
    totals = {"label": "Grand Total", "values": {"total": total}}
    subtitle = f"Period: {from_date} to {to_date}"

    if fmt == "excel":
        buf = build_excel("Daily Cash Book", columns, data, totals, subtitle)
        return excel_response(buf, f"daily_cashbook_{from_date}_to_{to_date}")
    else:
        buf = build_pdf("Daily Cash Book", columns, data, totals, subtitle, landscape_mode=False)
        return pdf_response(buf, f"daily_cashbook_{from_date}_to_{to_date}")


# ═══════════════════════════════════════════════════════════════
#  REPORT 5 — STUDENT FEE LEDGER (per-student full history)
# ═══════════════════════════════════════════════════════════════
def _ledger_data(db, cu, student_id, academic_year_id=None):
    params = {"tid": tid(cu), "sid": str(student_id)}
    extra = ""
    if academic_year_id:
        extra = " AND fi.academic_year_id = :yr"
        params["yr"] = str(academic_year_id)

    student = db.execute(text("""
        SELECT s.first_name, s.last_name, s.admission_no,
               g.name AS grade_name, sec.name AS section_name
        FROM students s
        LEFT JOIN student_enrollments se ON se.student_id=s.id AND se.status='active'
        LEFT JOIN sections sec ON sec.id = se.section_id
        LEFT JOIN grades g     ON g.id   = sec.grade_id
        WHERE s.id=:sid AND s.tenant_id=:tid
    """), params).fetchone()

    if not student:
        raise HTTPException(404, "Student not found")

    invoices = db.execute(text(f"""
        SELECT fi.invoice_no, fi.issue_date, fi.due_date, fi.total_amount,
               fi.paid_amount, fi.balance, fi.status
        FROM fee_invoices fi
        WHERE fi.student_id=:sid AND fi.tenant_id=:tid {extra}
        ORDER BY fi.issue_date
    """), params).fetchall()

    payments = db.execute(text(f"""
        SELECT fp.payment_date, fp.receipt_no, fp.method, fp.amount, fi.invoice_no
        FROM fee_payments fp
        JOIN fee_invoices fi ON fi.id = fp.invoice_id
        WHERE fi.student_id=:sid AND fp.tenant_id=:tid {extra}
        ORDER BY fp.payment_date
    """), params).fetchall()

    inv_data = [
        {
            "type": "Invoice",
            "date": str(r.issue_date),
            "ref":  r.invoice_no,
            "description": f"Due: {r.due_date}",
            "debit":  float(r.total_amount),
            "credit": 0,
            "status": r.status,
        }
        for r in invoices
    ]
    pay_data = [
        {
            "type": "Payment",
            "date": str(r.payment_date),
            "ref":  r.receipt_no or "—",
            "description": f"{(r.method or '').title()} — {r.invoice_no}",
            "debit":  0,
            "credit": float(r.amount),
            "status": "paid",
        }
        for r in payments
    ]
    combined = sorted(inv_data + pay_data, key=lambda x: x["date"])

    total_billed = sum(d["debit"] for d in combined)
    total_paid   = sum(d["credit"] for d in combined)

    return {
        "student": {
            "name": f"{student.first_name} {student.last_name or ''}".strip(),
            "admission_no": student.admission_no,
            "class": f"{student.grade_name or ''} {student.section_name or ''}".strip(),
        },
        "entries": combined,
        "total_billed": total_billed,
        "total_paid": total_paid,
        "balance": total_billed - total_paid,
    }


@router.get("/fees/ledger/{student_id}")
def student_ledger(
    student_id: UUID,
    academic_year_id: Optional[UUID] = Query(None),
    db: Session = Depends(get_db),
    cu: User    = Depends(get_current_user),
):
    return _ledger_data(db, cu, student_id, academic_year_id)


@router.get("/fees/ledger/{student_id}/export")
def export_ledger(
    student_id: UUID,
    fmt: str = Query("excel", regex="^(excel|pdf)$"),
    academic_year_id: Optional[UUID] = Query(None),
    db: Session = Depends(get_db),
    cu: User    = Depends(get_current_user),
):
    ledger = _ledger_data(db, cu, student_id, academic_year_id)

    rows = [
        {
            "date": e["date"],
            "type": e["type"],
            "ref":  e["ref"],
            "description": e["description"],
            "debit":  e["debit"] if e["debit"] > 0 else "",
            "credit": e["credit"] if e["credit"] > 0 else "",
        }
        for e in ledger["entries"]
    ]

    columns = [
        {"key":"date",        "label":"Date",        "width":12, "pdf_width":60},
        {"key":"type",        "label":"Type",        "width":10, "pdf_width":55},
        {"key":"ref",         "label":"Reference",   "width":15, "pdf_width":70},
        {"key":"description", "label":"Description", "width":25, "pdf_width":130},
        {"key":"debit",       "label":"Debit",       "width":12, "fmt":"currency", "pdf_width":65},
        {"key":"credit",      "label":"Credit",      "width":12, "fmt":"currency", "pdf_width":65},
    ]
    totals = {"label": "Balance", "values": {"debit": ledger["total_billed"], "credit": ledger["total_paid"]}}
    subtitle = (
        f"{ledger['student']['name']} ({ledger['student']['admission_no']}) — "
        f"{ledger['student']['class']}  |  Outstanding: Rs. {ledger['balance']:,.2f}"
    )

    title = f"Fee Ledger — {ledger['student']['name']}"
    if fmt == "excel":
        buf = build_excel(title, columns, rows, totals, subtitle)
        return excel_response(buf, f"ledger_{ledger['student']['admission_no']}")
    else:
        buf = build_pdf(title, columns, rows, totals, subtitle, landscape_mode=False)
        return pdf_response(buf, f"ledger_{ledger['student']['admission_no']}")


# ═══════════════════════════════════════════════════════════════
#  REPORT 6 — INCOME SUMMARY (category-wise totals, monthly)
# ═══════════════════════════════════════════════════════════════
def _income_summary_data(db, cu, academic_year_id):
    # Allocate each payment proportionally across its invoice's fee categories,
    # based on each category's share of the invoice subtotal.
    rows = db.execute(text("""
        WITH invoice_category_totals AS (
            SELECT
                fii.invoice_id,
                COALESCE(fc.name,'General') AS category,
                SUM(fii.amount)::float AS cat_amount,
                SUM(fii.amount) OVER (PARTITION BY fii.invoice_id)::float AS invoice_total
            FROM fee_invoice_items fii
            LEFT JOIN fee_categories fc ON fc.id = fii.fee_category_id
            GROUP BY fii.invoice_id, fc.name, fii.amount
        )
        SELECT
            TO_CHAR(fp.payment_date,'YYYY-MM')  AS month,
            TO_CHAR(fp.payment_date,'Mon YYYY') AS month_label,
            ict.category,
            SUM(
                fp.amount * (ict.cat_amount / NULLIF(ict.invoice_total, 0))
            )::float AS amount
        FROM fee_payments fp
        JOIN fee_invoices fi ON fi.id = fp.invoice_id
        JOIN invoice_category_totals ict ON ict.invoice_id = fi.id
        WHERE fp.tenant_id=:tid AND fi.academic_year_id=:yr
        GROUP BY TO_CHAR(fp.payment_date,'YYYY-MM'), TO_CHAR(fp.payment_date,'Mon YYYY'), ict.category
        ORDER BY month
    """), {"tid": tid(cu), "yr": str(academic_year_id)}).fetchall()

    # Pivot: month -> {category: amount}
    months = {}
    categories = set()
    for r in rows:
        if r.month not in months:
            months[r.month] = {"month": r.month_label}
        months[r.month][r.category] = r.amount
        categories.add(r.category)

    categories = sorted(categories)
    data = []
    for m in sorted(months.keys()):
        row = months[m]
        row["total"] = sum(row.get(c, 0) for c in categories)
        for c in categories:
            row.setdefault(c, 0)
        data.append(row)

    grand_total = sum(d["total"] for d in data)
    return data, categories, grand_total


@router.get("/fees/income-summary")
def income_summary_report(
    academic_year_id: UUID = Query(...),
    db: Session = Depends(get_db),
    cu: User    = Depends(get_current_user),
):
    data, categories, total = _income_summary_data(db, cu, academic_year_id)
    return {"months": data, "categories": categories, "grand_total": total}


@router.get("/fees/income-summary/export")
def export_income_summary(
    fmt: str = Query("excel", regex="^(excel|pdf)$"),
    academic_year_id: UUID = Query(...),
    db: Session = Depends(get_db),
    cu: User    = Depends(get_current_user),
):
    data, categories, total = _income_summary_data(db, cu, academic_year_id)

    columns = [{"key":"month", "label":"Month", "width":14, "pdf_width":70}]
    for c in categories:
        columns.append({"key":c, "label":c, "width":15, "fmt":"currency", "pdf_width":70})
    columns.append({"key":"total", "label":"Total", "width":15, "fmt":"currency", "pdf_width":75})

    totals_vals = {c: sum(d.get(c,0) for d in data) for c in categories}
    totals_vals["total"] = total
    totals = {"label": "Grand Total", "values": totals_vals}

    if fmt == "excel":
        buf = build_excel("Income Summary Report", columns, data, totals)
        return excel_response(buf, "income_summary")
    else:
        buf = build_pdf("Income Summary Report", columns, data, totals)
        return pdf_response(buf, "income_summary")
